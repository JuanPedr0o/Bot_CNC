from pathlib import Path
import os
from tkinter import filedialog, messagebox
from openpyxl.styles import Border, Side, Font
import openpyxl
import pandas as pd
from unidecode import unidecode
import sys
import shutil  # Adicione no topo com os outros imports

directory = Path(filedialog.askdirectory())

backup_dir = directory.parent / f"{directory.name}_BACKUP"  # Nome original + "_BACKUP" fora da pasta

shutil.copytree(directory, backup_dir, dirs_exist_ok=True)

# Salva o caminho da pasta escolhida em um arquivo de configuração

with open('config_dir.txt', 'w') as f:
    f.write(str(directory))  # Garanta que está salvando o caminho completo e correto
    
# Cria planilha do lote
planilha_lote = openpyxl.load_workbook('Lote.xlsx', read_only=False)
caminho_planilha_lote = f'{directory}' +'\\' + f'{directory.stem}' + '.xlsx'
planilha_lote.active.cell(row=2, column=1).value = directory.stem
planilha_lote.save(caminho_planilha_lote)

# Cria planilha dos retalhos e pecas do lote
planilha_lote = openpyxl.load_workbook('modelo_retalho.xlsx', read_only=False)
caminho_planilha_retalho_lote = f'{directory}' +'\\' + f'{directory.stem}' + '_retalhos_pecas' + '.xlsx'
planilha_lote.save(caminho_planilha_retalho_lote)

# Cria planilha dos retalhos
planilha_retalhos = openpyxl.load_workbook('LISTA-RETALHOS.xlsm', read_only=False, keep_vba=True)
# planilha_retalhos.create_sheet('retalhos_pecas')
caminho_planilha_retalhos = f'{directory}' +'\\' + f'{directory.stem}' + '_retalhos' + '.xlsm'
planilha_retalhos.save(caminho_planilha_retalhos)

# Define o estilo da borda
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Define o estilo da fonte
font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')

# Linha que inicia a inclusao das chapas inteiras
linha_chapas = 6

# Linha que inicia a inclusao dos retalhos
linha_retalhos = 2


def retalhos_pecas(caminho_txt):
    # Adiciona para cada retalho o nome das pecas que estao nele

    #Recupera o nome do retalho
    caminho_txt = Path(caminho_txt)
    retalho = caminho_txt.stem

    #Salva txt em um arquivo generico do excel
    df = pd.read_csv(caminho_txt, sep='|', encoding='latin-1')
    # caminho_excel = str(os.path.dirname(caminho_txt) + f'\\{retalho}' + '.xlsx')
    print(df)
    df.to_excel('excel_importado.xlsx')

    #Formata o arquivo importado
    wb = openpyxl.load_workbook('excel_importado.xlsx', read_only=False)
    ws = wb.active
    ws.delete_rows(1)
    ws.delete_cols(1, 3)
    ws.delete_cols(2)
    ws.delete_cols(3)
    ws.delete_cols(5, 10)
    ws.delete_cols(8, 20)

    mr = ws.max_row

    for i in range(1, mr+1):
        ped = ws.cell(row=i, column=3).value
        mod = ws.cell(row=i, column=2).value
        peca = ws.cell(row=i, column=1).value
        cod = ws.cell(row=i, column=4).value
        mat = ws.cell(row=i, column=5).value
        comp = ws.cell(row=i, column=6).value
        larg = ws.cell(row=i, column=7).value

        ws.cell(row=i, column=2).value = ped
        ws.cell(row=i, column=3).value = mod
        ws.cell(row=i, column=4).value = peca
        ws.cell(row=i, column=5).value = cod
        ws.cell(row=i, column=6).value = mat
        ws.cell(row=i, column=7).value = comp
        ws.cell(row=i, column=8).value = larg
        ws.cell(row=i, column=1).value = retalho

    ws.insert_rows(1)
    ws.cell(row=1, column=1).value = 'RETALHO'
    ws.cell(row=1, column=2).value = 'PEDIDO'
    ws.cell(row=1, column=3).value = 'MODULO'
    ws.cell(row=1, column=4).value = 'PECA'
    ws.cell(row=1, column=5).value = 'CODIGO'
    ws.cell(row=1, column=6).value = 'MATERIAL'
    ws.cell(row=1, column=7).value = 'COMPRIMENTO'
    ws.cell(row=1, column=8).value = 'LARGURA'

    mr = ws.max_row
    mc = ws.max_column

    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            ws.cell(row=i, column=j).border = thin_border
            ws.cell(row=i, column=j).font = font


    #ws.unmerge_cells('A1:F300')
    wb.save(str(os.path.dirname(caminho_txt) + f'\\{retalho}' + '.xlsx'))


    #Transfere para a planilha de retalhos do lote
    wbC = openpyxl.load_workbook(str(os.path.dirname(caminho_txt) + f'\\{retalho}' + '.xlsx'))
    wsC = wbC.active

    wb = openpyxl.load_workbook(caminho_planilha_retalho_lote, read_only=False)
    ws = wb.active
    mr_total = ws.max_row+1

    mr = wsC.max_row
    mc = wsC.max_column
    for i in range(2, mr + 1):
        for j in range(1, mc + 1):
            c = wsC.cell(row=i, column=j)
            ws.cell(row=mr_total, column=j).value = c.value
            ws.cell(row=mr_total, column=j).border = thin_border
        mr_total = mr_total +1

    wb.save(caminho_planilha_retalho_lote)


def planilha_restos(nome_retalho):
    # Passa o codigo do retalho para a planilha de retalho

    # print('adicionando retalho na planilha' + nome_retalho)

    # Carrega planilha de retalhos
    wb = openpyxl.load_workbook(caminho_planilha_retalhos, read_only=False, keep_vba=True)
    ws = wb.active

    ws.cell(row=linha_retalhos, column=1).value = '*' + nome_retalho + '*'

    wb.save(caminho_planilha_retalhos)


def planilha_lote(nome_chapa):
    # Passa a chapa para a plainha de lote

    # Carrega planilha do lote
    wb = openpyxl.load_workbook(caminho_planilha_lote, read_only=False)
    ws = wb.active

    ws.cell(row=linha_chapas, column=1).value = nome_chapa
    ws.cell(row=linha_chapas, column=2).value = nome_chapa

    coluna = 1
    while coluna <= 4:
        ws.cell(row=linha_chapas, column=coluna).border = thin_border
        coluna += 1

    wb.save(caminho_planilha_lote)


def arruma_ordem(caminho_arquivo):

    arquivo2 = caminho_arquivo

    with open(arquivo2, 'r', encoding="latin-1") as etiquetas:
        linhas2 = etiquetas.readlines()
        primeira_linha = linhas2[0]
        linhas2.pop(0)

    linhas2.sort()

    with open(arquivo2, 'w', encoding="latin-1") as etiquetas:
        etiquetas.write(primeira_linha)

    with open(arquivo2, 'a', encoding="latin-1") as etiquetas:
        for linha2 in linhas2:
            etiquetas.write(str(linha2))


for (root, dirs, files) in os.walk(directory, topdown=True):
    for file in files:
        arquivo = Path(file)

        if arquivo.name.endswith('.TXT'):

            caminho_completo = os.path.join(root, file)
            arruma_ordem(caminho_completo)

for (root, dirs, files) in os.walk(directory, topdown=True):
    for file in files:
        arquivo = Path(file)

        if arquivo.name.endswith('.xxl'): #########################################################
            with open(os.path.join(root, file), 'r', errors='ignore') as xxl:

                linhas = xxl.readlines()
                total_linhas = len(linhas)
                print(total_linhas)
            i = 0
            encontrado = 0
            while i < total_linhas:
                if linhas[i].find('Load *') != -1:
                    ini = linhas[i].find('Load *')
                    ini2 = ini + 6
                    fim = linhas[i].find('* Dim')
                    linha = str(linhas[i])
                    cod_xxl = linha[ini2:fim]+'.xxl' ######################################
                    cod_txt = linha[ini2:fim]+'.TXT'

                    renomear_xxl = os.path.join(root, cod_xxl)
                    print(renomear_xxl)
                    caminho1_txt = str(os.path.join(root, file))[:-4]
                    caminho2_txt = caminho1_txt + '.TXT'
                    renomear_txt = os.path.join(root, cod_txt)

                    os.rename(os.path.join(root, file), renomear_xxl)
                    os.rename(caminho2_txt, renomear_txt)

                    planilha_restos(str(linha[ini2:fim]))
                    linha_retalhos = linha_retalhos + 1

                    retalhos_pecas(renomear_txt)

                    encontrado = 1
                i += 1

            if encontrado == 0:
                planilha_lote(arquivo.stem)
                linha_chapas += 1


# --- Nova funcionalidade para processar pastas "QUALQUER COR" ---
def processar_qualquer_cor(diretorio_principal):
    caminho_excel = directory / "Relatorio_Qualquer_Cor.xlsx"
    
    # Criar o ExcelWriter e garantir que a planilha 'Dados' exista
    writer = pd.ExcelWriter(caminho_excel, engine='openpyxl')
    workbook = writer.book
    
    # Criar a planilha 'Dados' se não existir
    if 'Dados' not in workbook.sheetnames:
        worksheet = workbook.create_sheet(title='Dados')
    else:
        worksheet = workbook['Dados']
    
    # Remover a planilha padrão 'Sheet' (se houver)
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    worksheet.title = "Dados"
    linha_atual = 1  # Iniciar na linha 1
    
    # Procurar pastas "QUALQUER COR"
    for root, dirs, files in os.walk(diretorio_principal):
        for dir_name in dirs:
            if 'QUALQUER COR' in dir_name.upper():
                pasta_cor = Path(root) / dir_name
                
                # Processar cada TXT na pasta
                for arquivo_txt in pasta_cor.glob('*.TXT'):
                    df_temp = pd.DataFrame(columns=['Código', 'Material', 'Espessura', 'Comprimento', 'Largura'])  # Inicialização garantida
                    
                    # Adicionar título do arquivo
                    worksheet.cell(row=linha_atual, column=1, value=f"Arquivo: {arquivo_txt.name}")
                    worksheet.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=5)
                    linha_atual += 1
                    
                    # Ler e processar o TXT
                    with open(arquivo_txt, 'r', encoding='latin-1') as f:
                        for linha in f:
                            dados = linha.strip().split('|')
                            if len(dados) < 22:
                                continue
                            try:
                                codigo = dados[8].strip()
                                material = dados[19].strip()
                                espessura = material.split()[-1].replace('mm', '').strip()
                                comprimento = dados[20].strip()
                                largura = dados[21].strip()
                                
                                novo_registro = pd.DataFrame([{
                                    'Código': codigo,
                                    'Material': material,
                                    'Espessura': espessura,
                                    'Comprimento': comprimento,
                                    'Largura': largura
                                }])
                                
                                df_temp = pd.concat([df_temp, novo_registro], ignore_index=True)
                                
                            except Exception as e:
                                print(f"Erro no arquivo {arquivo_txt.name}: {e}")
                                continue

                    # Escrever dados no Excel (APENAS SE df_temp NÃO ESTIVER VAZIO)
                    if not df_temp.empty:
                        # Escrever cabeçalhos
                        for c_idx, col_name in enumerate(df_temp.columns, start=1):
                            worksheet.cell(row=linha_atual, column=c_idx, value=col_name)
                        linha_atual += 1
                        
                        # Escrever dados
                        for _, row in df_temp.iterrows():
                            for c_idx, value in enumerate(row, start=1):
                                worksheet.cell(row=linha_atual, column=c_idx, value=value)
                            linha_atual += 1
                        
                        linha_atual += 1  # Pular uma linha após cada bloco

    # Formatar o Excel
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.font = font

    # Ajustar largura das colunas (CORRIGIDO)
    for col_idx in range(1, worksheet.max_column + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # Ignorar células mescladas
        for cell in worksheet[column_letter]:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    writer.close()

# Executar antes do messagebox final
processar_qualquer_cor(directory)
    

# messagebox.showinfo(title='Processos Finalizados', message='Todos os processos foram finalizados. Iniciando scripts secundários...')

# Executar scripts na ordem correta, usando o mesmo diretório
scripts = ["GerarNormais.py", "GerarOsDe25.py", "GerarRetalhos.py", "Extras.py"]
for script in scripts:
    try:
        os.system(f'python "{script}"')  # Assume que os scripts estão no mesmo diretório
    except Exception as e:
        print(f"Erro ao executar {script}: {e}")
